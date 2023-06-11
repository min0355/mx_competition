import os
import sys
import re
import pandas as pd

from PySide6.QtCore import Qt, QCoreApplication, QItemSelectionModel, QTimer
from PySide6.QtWidgets import (QApplication, QFileDialog, QHeaderView, QHBoxLayout, 
                               QLineEdit, QPushButton, QTableWidget, QTableWidgetItem, 
                               QTreeWidget, QTreeWidgetItem, QVBoxLayout, QWidget,
                               QDialog, QDialogButtonBox, QFormLayout, QLabel, QVBoxLayout, 
                               QLineEdit, QLabel, QMessageBox, QGroupBox, QAbstractItemView)
from PySide6.QtGui import QColor, QFont
from PySide6 import QtGui
import string
from collections import OrderedDict

import openpyxl
from openpyxl.styles import PatternFill


'''

FINISHED

- ape 삭제 -> 완료 
- 실행 시 전체 창, 품번이 다 보이는 열 너비로 default 세팅 -> 완료 
- 프로그램 실행 창 제목 입력 -> 완료
- 삭제 품번 체크 반영되도록 수정 -> 완료 
- 품번, PREFIX, 수량, 단위는 정확히 입력 필요하고 품명은 대충 넣어도 되도록 -> bom 완성도를 위해선 정확히 넣는걸로
- 엑셀 저장 -> 완료
- 버튼 순서 및 이름 수정 ->  완료
- excel save 시 parent 정보 추가 -> 완료
- 최종 테이블 화면 색상 표시 (추가, 변경) -> 완료
- PARENT 가 무첨자일때 A 로 변경 되지 않는 문제 -> 완료
- LP 소개 추가 -> 완료 
- 상단 버튼 최종 정리 (버튼 종류별로 색상 구분되면 편할 듯, 서치 버튼 좀 더 넓게 => 완료)  
- 상단 버튼 tap 2 줄 -> 완료 
- 11 자리 품번 외엔 PARENT 자동 첨자 변경이 되지 않음  -> 완료
- 첨자가 설변 품번 개수만큼 올라가는 것 수정 (신규 품번 추가 + 설변 품번, 삭제 품번 모두 첨자는 1 번만 변경되어야 함.) -> 완료
- 신규품번 삭제품번 I,D 네이밍 변경 -> 완료 
- class -> itm 으로 수정 -> 완료
- Bom comparison 창 최소화 기능 추가 -> 완료 
- Tree 구조 버튼 없이 엑셀 불러오면 둘 다 자동으로 출력되도록 수정 -> 완료
- 생성해야 할 품번 리스트를 엑셀 우측에 정렬 (창 삭제) + INSERT 품번도 같이 보여주기 -> (정홍연/5.29완료) 
- 상단 버튼 재정렬
- 불 필요 행 삭제 혹은 삭제 할 품번 체크 시 알림창 추가 (Undo 기능 대체, Undo 기능 구현하지 않기로 결정) -> (정홍연/5.29완료)
- 불 필요 행 삭제 & 삭제 품번 체크 -> 두 개 버튼의 의미가 헷갈려서 유저 편의성 개선 필요 (버튼 재정렬 시 색상으로 구분하거나 ) -> 완료
- 초기 display 둘 다 트리 구조로 디스플레이 -> (정홍연/5.29완료)
- 리턴 기능 -> 삭제 버튼 알람으로 대체 -> 완료
- 삭제 버튼만 입력 시에도 change name 버튼 동작 되도록 수정 (+ 신규 품번 추가 등을 취소하거나 삭제 품번 체크 먼저 하면 기능 동작 안 함, 프로그램 실행 후 바로 실행하면 문제 없이 작동) -> 완료
- "Transform to Table" 버튼 2회째부터는 스크롤 링크, 색상 변경이 사라짐 -> 완료
- ASSY 품번 추가 기능(팝업) -> 완료
- SELF.QTREE의 정체 -> 완료
- delete복구 기능 추가 -> 완료 
- delete된 품번에 신규 품번 생성 막음 -> 완료 
- delete 품번의 하위 품번 자동 삭제 -> 완료 
- EXE 파일로 만드는 방법 (PYTHON 미 설치에도 쓸 수 있도록) -> 완료
- 테이블, 트리 사이즈 조정 -> 완료
- 엑셀 로딩 시 트리 최상위 행에 해당하는 테이블 빈행으로 만듦(DF x) -> 완료
- 신규품번 창생 시 최상위 노드를 제외한 순번 카운팅 및 DF수정, 리프레쉬(DF o) -> 완료
- 트리 2 테이블 변환(트랜스폼 창)에서 신규 창생 품번 빈행으로 표기되게 함 -> 완료
- '신규 품번 추가(I)' 시 NEW BOM 선택시에만 작동, PARENT 첨자 update 작동 기능 없음 -> 완료
- 왼쪽 삭제 버튼 삭제 -> 완료
- Old bom display 창에 New bom과 같이 보기 편하게 blank row 추가  -> 완료
- NEW BOM의 PARENT 첨자 update 시 OLDBOM과 동일 PARENT 탐색 불가 --> OLD BOM에 BLANK 추가 시 동일 PARENT 탐색 필요 (손경호)
  --> 좌측 테이블 구조로 변경하면서 이슈 소멸
- CHANGE NAME 을 CHANGE ITM 으로 변경 -> 완료
- CHANGE NAME 버튼을 신 품번 적용으로 변경 -> 완료
- CHANGE NAME 의 칸 2 개 사이에 화살표 추가하면 좋음 (구 -> 신품번) -> 완료 
- SRC, PROC 는 전산상 자동 입력이니 굳이 PROPERTIES 에 넣을 이유 없어 보임 -> 전체 수정의 의미가 있으므로 일단 그대로 두자, 완료
- 변경 품번을 여러개 동시에 바꾸는 것 필요 여부? -> voc 정도 접수로 처리
- '삭제 품번 체크(D)' 또한 PARENT 첨자 UPDATE 작동 기능 없음 -> 완료 
- 결과값 엑셀 열 정렬 -> 완료
- search 버튼 이후 품번 변경 버튼 동작하지 않음 -> 완료
- 글씨체 동일 (글씨체 다른 것이 나을 수 있음.) -> 완료 
- 오른쪽 트리에 줄 표시 (왼쪽 처럼) -> 줄 표시 없어도 무방, 완료 
- E 하위 밑에 F 가 없는 경우 (ASSY 는 있어야 되는데, 스핀들은 통상 없음. 두 가지 타입 모두 존재) -> 사용자에게 맡김. 
- - 삭제 버튼 동작 하지 않음 (지워지지 않음) -> 완료 


~ TO DO LIST 


손경호/정홍연
- 삭제, 삽입 된 행도 고려하여 저장된 엑셀 파일을 그대로 복붙 할 수 있도록 최종 저장 BOM 구성
- Old bom & New bom 디스플레이 시 flag 추가 



김영민

- '삭제 복구 버튼'이 정상 설변 (RED) 된 행들도 색상만 다시 돌리는 오류 (바뀐 첨자는 유지), 삭제 체크 복구 시 설변 (빨강) 이 원복
- BOM 내 동일 품번들 중 일부만 바뀌는 경우에 대한 알고리즘 

- 공통된 설변에서 한 OPTN 은 설변하고 다른 OPTN 은 그대로 유지하기 위한 설변 롤백 버튼 추가 필요 


- 세련된 버튼 적용 (초기화 버튼 추가 필요해 보임. 모든 케이스에서 작동이 엉킬경우 초기화 버튼 필요 (다시 껐다 실행하거나))



~ TEST 


- 최종 확인 절차 BOM (OPTN 전부 확인 가능) -> 최종 BOM 을 믿을 수 있느냐는 의문이고, 테스트 통해 최종 배포 전 확인 필요 



~ MONITORING

- - PREFIX 중복이 되는 경우 (MOVE UP / DOWN) 원래 PREFIX 색상 칠해지면 좋음. (중복 확인 용이하도록)
- Move up&down 버튼 구현할 지 말지 결정 필요-> 배포 후 의견 수렴
- 정규식 추가 (품번 잘못 입력 시 알림 창)  -> 완료 (모니터링 필요)
  
'''

QCoreApplication.setAttribute(Qt.AA_EnableHighDpiScaling, True)

def tree_to_dataframe(tree_widget):
    data = []
    
    def visit_node(node, parent=None, level=0):
        if parent is not None:
            row = [level, parent]
            row.append(node.text(1))  # Add PREFIX

            # Add text and background color for each column
            for j in range(tree_widget.columnCount()):
                bg_color = node.background(j).color()
                if bg_color.isValid():  # Check if the background color is valid
                    if bg_color == QColor.fromRgbF(0, 0, 0, 1):  # Check if the background color is black
                        row.append({'text': node.text(j), 'bg_color': Qt.white})  # Use white as the background color
                    else:
                        row.append({'text': node.text(j), 'bg_color': bg_color})
                else:
                    row.append({'text': node.text(j), 'bg_color': Qt.white})  # Use white as the default background color
                
            row.pop(4)  # Remove CLASS for ITM
            data.append(row)

        for i in range(node.childCount()):
            visit_node(node.child(i), node.text(0), level + 1)

    for i in range(tree_widget.topLevelItemCount()):
        item = tree_widget.topLevelItem(i)
        visit_node(item)

    columns = ["LVL", "PARENT", "PREFIX", "ITM", "ITM_DESC", "QTY", "UOM", "SRC", "PROC", "THREAD"]
    df = pd.DataFrame(data, columns=columns)
    return df


class MyTreeWidget(QTreeWidget):
     
    def __init__(self, parent=None):
        super(MyTreeWidget, self).__init__(parent)
        self.setUniformRowHeights(True)  


        self.setHeaderLabels(["ITM", "PREFIX", "ITM_DESC", "QTY", "UOM", "SRC", "PROC", "THREAD"])
        self.setAlternatingRowColors(True)
        self.header().setSectionResizeMode(QHeaderView.Interactive)
        self.itemClicked.connect(self.item_clicked)


    def do_search(self, term):
        clist = self.findItems(term, Qt.MatchContains | Qt.MatchRecursive)

        sel = self.selectionModel()
        for item in clist:
            index = self.indexFromItem(item)
            sel.select(index, QItemSelectionModel.Select)
        
    def set_item_background_color(self, item, color):
        for i in range(self.columnCount()):
            item.setBackground(i, color)

    def item_clicked(self, item, column):
       self.parent().parent().load_item_properties(item)
       
    def count_all_nodes(self, item):
        count = 1  # count the item itself
        for i in range(item.childCount()):
            count += self.count_all_nodes(item.child(i))
        return count

    def contains_item(self, root, target):
        # Check if root is the target
        if root == target:
            return True
        # Recursively check all children
        for i in range(root.childCount()):
            if self.contains_item(root.child(i), target):
                return True
        # If the loop finishes, the item is not in this subtree
        return False

    def calculate_row_index(self, selected_item, root_item=None, index=0):
        if root_item is None:
            root_item = self.invisibleRootItem()

        for i in range(root_item.childCount()):
            child_item = root_item.child(i)
            
            # if child_item.parent() is not None:
            if self.contains_item(child_item, selected_item):
                if child_item == selected_item:
                    return index + 1
                else:
                    if child_item.parent() is None:
                        return self.calculate_row_index(selected_item, child_item, index)
                    else:
                        return self.calculate_row_index(selected_item, child_item, index + 1)
            else:
                if child_item.parent() is None:
                    index += self.count_all_nodes(child_item) 
                    index = index -1
                else: 
                    index += self.count_all_nodes(child_item) 
        return index  



class DataFrameDialog(QDialog):
    def __init__(self, parent=None):
        super(DataFrameDialog, self).__init__(parent)

        self.setWindowTitle('BOM Comparison')
        self.setWindowFlags(self.windowFlags() | Qt.WindowMaximizeButtonHint | Qt.WindowMinimizeButtonHint) 

        self.left_table = QTableWidget()
        self.right_table = QTableWidget()

        layout = QVBoxLayout()

        left_table_title = QLabel("OLD BOM")
        right_table_title = QLabel("NEW BOM")

        title_layout = QHBoxLayout()
        title_layout.addWidget(left_table_title)
        title_layout.addWidget(right_table_title)
        
        layout.addLayout(title_layout)

        
        table_layout = QHBoxLayout()
        table_layout.addWidget(self.left_table)
        table_layout.addWidget(self.right_table)
        
        layout.addLayout(table_layout)
        
        
        self.setLayout(layout)




    def show_dataframes(self, old_df, new_df):
        self.show_dataframe(self.left_table, old_df)
        self.show_dataframe(self.right_table, new_df)
        self.compare_and_highlight_differences(old_df, new_df)
        self.sync_table_scrollbars()
        self.show()

    def show_dataframe(self, table, df):
        table.setRowCount(df.shape[0])
        table.setColumnCount(df.shape[1])
        table.setHorizontalHeaderLabels(df.columns)

        for i in range(df.shape[0]):
            for j in range(df.shape[1]):
                cell_data = df.iat[i, j]
                if isinstance(cell_data, dict):  # Check if cell_data is a dictionary
                    item = QTableWidgetItem(cell_data['text'])
                    item.setBackground(cell_data['bg_color'])
                else:
                    item = QTableWidgetItem(str(cell_data))                    
                table.setItem(i, j, item)

    def compare_and_highlight_differences(self, old_df, new_df):
        # Create a set of row labels for each dataframe
        old_row_labels = set(old_df.index)
        new_row_labels = set(new_df.index)

        # Find missing rows in old_df
        missing_rows = new_row_labels - old_row_labels

        # Insert missing rows into the old_df table with a blue background
        for row_label in missing_rows:
            new_row_index = new_df.index.get_loc(row_label)
            self.left_table.insertRow(new_row_index)
            for j in range(old_df.shape[1]):
                item = QTableWidgetItem("")
                item.setBackground(Qt.yellow)
                self.left_table.setItem(new_row_index, j, item)
                
    def sync_table_scrollbars(self):
        def sync_scrollbars(value):
            self.left_table.verticalScrollBar().setValue(value)
            self.right_table.verticalScrollBar().setValue(value)

        left_scrollbar = self.left_table.verticalScrollBar()
        right_scrollbar = self.right_table.verticalScrollBar()

        left_scrollbar.valueChanged.connect(sync_scrollbars)
        right_scrollbar.valueChanged.connect(sync_scrollbars)   

class AddRowDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
 
        self.setWindowTitle("Add New Row")

        vbox = QVBoxLayout()

        form_layout = QFormLayout()
        self.class_edit = QLineEdit()
        self.prefix_edit = QLineEdit()
        self.itm_desc_edit = QLineEdit()
        self.qty_edit = QLineEdit()
        self.uom_edit = QLineEdit()

        form_layout.addRow("ITM:", self.class_edit)
        form_layout.addRow("Prefix:", self.prefix_edit)
        form_layout.addRow("ITM_DESC:", self.itm_desc_edit)
        form_layout.addRow("QTY:", self.qty_edit)
        form_layout.addRow("UOM:", self.uom_edit)

        vbox.addLayout(form_layout)

        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        button_box.accepted.connect(self.accept)
        button_box.rejected.connect(self.reject)

        vbox.addWidget(button_box)

        self.setLayout(vbox)


    # V2 정규식  

    def add_error_check(self):
        error_messages = []
        
        rule_1 = r'(^\d{6}-\d{5}[A-Za-z]?)|(R\d{5}[A-Za-z]?$)|(r\d{5}[A-Za-z]?$)|(\d{6})|(^C\d{8}[A-Za-z]?$)|(^c\d{8}[A-Za-z]?$)|(^L\d{8}[A-Za-z]?$)|(^l\d{8}[A-Za-z]?$)|(^S\d{7})|(^s\d{7})|(^P\d{8})|(^p\d{8})|(^\d{6}-\d{5}[A-Za-z]?$-M)|(^\d{6}-\d{5}[A-Za-z]?$-m)|(^EBUSH\d{4})|(^ebush\d{4})|^R\d{6}[A-Za-z]?$|^r\d{6}[A-Za-z]?$|(^MVZP\d{3})|(^mvzp\d{3})|(^EHOSE\d{4}[A-Za-z]?$)|(^ehose\d{4}[A-Za-z]?$)|(^ETUBN\d{4})|(^etubn\d{4})'
        rule_2 = r'(^MVZH\d{3}-\d{3}$)|(^\d{6}[A-Za-z]?$)|(^L\d{8}[A-Za-z]?-\d{3}$)|(^L[A-Za-z]{3}\d{3}$|^L[A-Za-z]{4}\d{2}[A-Za-z]$|^L[A-Za-z]{3}\d{3}[A-Za-z]$)|(^L[A-Za-z]{3}\d{3}[A-Za-z]\d{3}$)|(^A[A-Za-z]{2}\d{3}A\d$)|(^YP\d{7}$|^YX\d{7}$)|(^T\d{7}[A-Za-z]?$)|(^E[A-Za-z]{4}\d{4}[A-Za-z]?$)'

        if not re.match(rule_1, self.class_edit.text()) and not re.match(rule_2, self.class_edit.text()):
            error_messages.append('품번 입력 규칙에 맞지 않습니다.')

        if not self.prefix_edit.text().isdigit() or len(self.prefix_edit.text()) != 4:
            error_messages.append('Prefix는 4자리 숫자만 입력 가능합니다.')

        if not self.qty_edit.text().isdigit():
            error_messages.append('수량은 숫자만 입력 가능합니다.')

        return error_messages



    def get_row_data(self):
        self.add_error_check()
       
        return [self.class_edit.text(), self.prefix_edit.text(), self.itm_desc_edit.text(),
                self.qty_edit.text(), self.uom_edit.text()]
    
    # 추가한 부분 (위)
    '''
    def do_search(self, term):
        clist = self.findItems(term, Qt.MatchContains | Qt.MatchRecursive)

        sel = self.selectionModel()
        for item in clist:
            index = self.indexFromItem(item)
            sel.select(index, QItemSelectionModel.Select)
            
            
    def set_item_background_color(self, item, color):
        for i in range(self.columnCount()):
            item.setBackground(i, color)

    def item_clicked(self, item, column):
        self.parent().load_item_properties(item)
    '''
    # 추가한 부분 (아래)
 
class myWindow(QWidget):
    def __init__(self):
        super().__init__()
        
        self.setWindowTitle("AUTO BOM (Mady by Ars Machina)")

        
        self.df_list = []
        self.old_names = []
        self.updated_names = []

        # V5
        self.deleted_items = []

        open_btn = QPushButton('엑셀 파일 열기', self)
        self.line_edit = QLineEdit()
        self.line_edit.setFixedWidth(100) 
        
        search_button = QPushButton("Search")
        transform_btn = QPushButton('Transform to Table', self)
          
        tree_move_up_btn = QPushButton('Move Up', self)
        tree_move_down_btn = QPushButton('Move Down', self)

        tree_add_btn = QPushButton('신규 품번 추가 (I)', self)
        remove_btn = QPushButton('삭제 품번 체크 (D)', self)      
        recover_btn = QPushButton('삭제 체크 복구', self)    
     
 
        # 수평 박스 배치
        hbox = QHBoxLayout()
        hbox.addWidget(open_btn)

        hbox.addWidget(self.line_edit)
        hbox.addWidget(search_button)
        hbox.addWidget(tree_move_up_btn)
        hbox.addWidget(tree_move_down_btn)   
  
        hbox.addWidget(tree_add_btn)
        hbox.addWidget(remove_btn)
        hbox.addWidget(recover_btn)
        hbox.addWidget(transform_btn)        
                
        # self.qtree_original = MyTreeWidget(self)
        
        self.table = QTableWidget(self)
        self.qtree = MyTreeWidget(self)
        
        # Creating QGroupBox for each widget
        group_box1 = QGroupBox("Old Bom")
        group_box1.setLayout(QVBoxLayout())
        group_box1.layout().addWidget(self.table)
        group_box1.setStyleSheet("""
            QGroupBox {
                font-size: 16px;
                margin-top: 10px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 400%;
            }
        """)

        group_box2 = QGroupBox("New Bom")
        group_box2.setLayout(QVBoxLayout())
        group_box2.layout().addWidget(self.qtree)
        group_box2.setStyleSheet("""
            QGroupBox {
                font-size: 16px;
                margin-top: 10px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 400%;
            }
        """)

        # Adding the group boxes to TableTreeBox layout
        TableTreeBox = QHBoxLayout()
        TableTreeBox.addWidget(group_box1)
        TableTreeBox.addWidget(group_box2)
        
        change_name_label = QLabel("Change ITM:")
        self.old_name_edit = QLineEdit()
        self.new_name_edit = QLineEdit()
        self.old_name_edit.setFixedWidth(100)  
        self.new_name_edit.setFixedWidth(100)  
        self.change_name_button = QPushButton("신 품번 적용")

        # V5

        arrow_label = QLabel("→", self)
        arrow_label.setStyleSheet("font-size: 20px; font-weight: bold; color: blue;")


        right_hbox = QHBoxLayout()
        right_hbox.addWidget(change_name_label)
        right_hbox.addWidget(self.old_name_edit)
        right_hbox.addWidget(arrow_label)
        right_hbox.addWidget(self.new_name_edit)
        right_hbox.addWidget(self.change_name_button)

        self.property_labels = ["ITM", "PREFIX", "ITM_DESC", "QTY", "UOM", "SRC", "PROC", "THREAD"]
        self.property_lineedits = [QLineEdit() for _ in range(len(self.property_labels))]
        properties_layout = QHBoxLayout()
        for label, lineedit in zip(self.property_labels, self.property_lineedits):
            properties_layout.addWidget(QLabel(label))
            properties_layout.addWidget(lineedit)

        self.save_properties_btn = QPushButton("Save Properties")
        self.save_properties_btn.clicked.connect(self.save_item_properties)
        properties_layout.addWidget(self.save_properties_btn)

        right_hbox.addLayout(properties_layout)
        
        # 수직 박스 배치
        vbox = QVBoxLayout()
        vbox.addLayout(hbox)
        vbox.addLayout(right_hbox)
        vbox.addLayout(TableTreeBox)
        
        self.setLayout(vbox)

        # 시그널 연결
        open_btn.clicked.connect(self.clickOpenBtn)
        search_button.clicked.connect(self.on_search_button_clicked)
        self.change_name_button.clicked.connect(self.on_change_name_button_clicked)
        
        tree_add_btn.clicked.connect(self.clickTreeAddBtn)
        tree_move_up_btn.clicked.connect(self.clickTreeMoveUpBtn)
        tree_move_down_btn.clicked.connect(self.clickTreeMoveDownBtn)
        
        
        remove_btn.clicked.connect(self.clickRemoveBtn)
        recover_btn.clicked.connect(self.clickRecoverBtn)
        transform_btn.clicked.connect(self.on_transform_button_clicked)

        # Change background color of buttons and input windows
        change_name_color = "background-color: #FDEBD0;"

        self.old_name_edit.setStyleSheet(change_name_color)
        self.new_name_edit.setStyleSheet(change_name_color)
        self.change_name_button.setStyleSheet(change_name_color)

        for lineedit in self.property_lineedits:
            lineedit.setStyleSheet(change_name_color)

        self.save_properties_btn.setStyleSheet(change_name_color)


        tree_add_btn.setStyleSheet("background-color: #ADD8E6;")
        remove_btn.setStyleSheet("background-color: #ADD8E6;")
        
        # scrollbars
        self.table.verticalScrollBar().valueChanged.connect(self.sync_scroll)
        self.qtree.verticalScrollBar().valueChanged.connect(self.sync_scroll)


    def sync_scroll(self, value):
            sender = self.sender()
            if sender == self.table.verticalScrollBar():
                self.qtree.verticalScrollBar().setValue(value) 
            elif sender == self.qtree.verticalScrollBar():
                self.table.verticalScrollBar().setValue(value)       


    def clickTreeAddBtn(self):
        selected_item = self.qtree.currentItem()

        if not selected_item:
            return

        item_name = selected_item.text(0)
        
        RowIndex_Count = self.qtree.calculate_row_index(selected_item)
        
        data = selected_item.data(0, Qt.UserRole)
        if data == "deleted":
            message_box = QMessageBox()
            message_box.setText(f"{item_name}의 데이터는 'deleted'로 설정되었습니다.")
            message_box.exec()
            return

        reply = QMessageBox.question(self, 'Message',
                                    f"{item_name}의 하위품번으로 등록하시겠습니까?", QMessageBox.Yes |
                                    QMessageBox.No, QMessageBox.No)
        
        if reply == QMessageBox.Yes:
            parent_item = selected_item
        else:
            parent_item = selected_item.parent() if selected_item.parent() else self.qtree

        dialog = AddRowDialog(self)
        result = dialog.exec()

        if result == QDialog.Accepted:
            error_messages = dialog.add_error_check()            
            if error_messages:
                parent_item = []
                self.show_alert('\n'.join(error_messages))
                self.clickTreeAddBtn()
            else:
                new_item_data = dialog.get_row_data()
                new_item = QTreeWidgetItem(new_item_data)
            
                self.updated_names.append(new_item_data[0])

                yellow_background = QColor(255, 255, 0)
                for i in range(new_item.columnCount()):
                    new_item.setBackground(i, yellow_background)
            
                parent_item.insertChild(parent_item.indexOfChild(selected_item) + 1, new_item)
                
                new_row = pd.DataFrame([['' for _ in range(self.df_list[2].shape[1])]], columns=self.df_list[2].columns)
                self.df_list[2] = pd.concat([self.df_list[2].iloc[:RowIndex_Count], new_row, self.df_list[2].iloc[RowIndex_Count:]]).reset_index(drop=True)
                
                # Update the table display
                self.initTableWidget(2)

                # V5 
                self.change_node_and_ancestors(parent_item)
                    
    



            
    def clickTreeMoveUpBtn(self):
        selected_item = self.qtree.currentItem()
        if not selected_item or not selected_item.parent():
            return

        parent = selected_item.parent()
        index = parent.indexOfChild(selected_item)
        if index == 0:
            return

        parent.removeChild(selected_item)
        parent.insertChild(index - 1, selected_item)
        self.qtree.setCurrentItem(selected_item)

    def clickTreeMoveDownBtn(self):
        selected_item = self.qtree.currentItem()
        if not selected_item or not selected_item.parent():
            return

        parent = selected_item.parent()
        index = parent.indexOfChild(selected_item)
        if index == parent.childCount() - 1:
            return

        parent.removeChild(selected_item)
        parent.insertChild(index + 1, selected_item)
        self.qtree.setCurrentItem(selected_item)


    def clickRemoveBtn(self):
        selected_items = self.qtree.selectedItems()
        for item in selected_items:
            item_name = item.text(0)

        reply = QMessageBox.question(self, 'Message',
                                    item_name + "및 하위품번을 BOM에서 삭제 하시겠습니까?", QMessageBox.Yes |
                                    QMessageBox.No, QMessageBox.No)
        if reply == QMessageBox.Yes:
            gray_background = QColor(192, 192, 192)
            for item in selected_items:
                item.setData(0, Qt.UserRole, "deleted")
                self.CheckChild_change_tag_color(item, Qt.UserRole, "deleted", gray_background)
                # v5
                self.deleted_items.append(item.text(0))  # 삭제 리스트를 추가 
                self.change_node_and_ancestors(item.parent())
        else:
            None


    def CheckChild_change_tag_color(self, item, role, value, color):
        for i in range(item.columnCount()):
            item.setBackground(i, color)

        child_count = item.childCount()
        for i in range(child_count):
            child_item = item.child(i)
            child_item.setData(0, role, value)
            self.CheckChild_change_tag_color(child_item, role, value, color)

    
    def clickRecoverBtn(self):
        selected_items = self.qtree.selectedItems()
        for item in selected_items:
            item_name = item.text(0)

        reply = QMessageBox.question(self, 'Message', "삭제된 품번을 복구 하시겠습니까?", QMessageBox.Yes |
                                    QMessageBox.No, QMessageBox.No)
        if reply == QMessageBox.Yes:
            for item in selected_items:
                item.setData(0, Qt.UserRole, "")
                for i in range(item.columnCount()):
                    item.setBackground(i, Qt.white)
        else:
            None


    
              
    def on_transform_button_clicked(self):
        df = tree_to_dataframe(self.qtree)
        if len(self.df_list) < 4:
            self.df_list.append(df)
        else:
            self.df_list[3] = df

        
        old_df = self.df_list[2]
        new_df = self.df_list[3]
        # self.show_dataframe_on_table(self.df_list[-1])
        self.show_dataframe_in_popup(old_df, new_df) 

        for column in new_df.columns:
            new_df[column] = new_df[column].apply(lambda x: x['text'] if isinstance(x, dict) and 'text' in x else x)

        self.save_dataframe_to_excel(new_df)    # 엑셀 신규 bom 만 저장 

    def show_dataframe_in_popup(self, old_df, new_df):
        dialog = DataFrameDialog(self)
        dialog.show_dataframes(old_df, new_df)
        dialog.setGeometry(100, 100, 1200, 800)
        dialog.exec()

    def load_item_properties(self, item):
        for i in range(len(self.property_labels)):
            self.property_lineedits[i].setText(item.text(i))
        self.current_item = item

    def save_item_properties(self):
        for i in range(len(self.property_labels)):
            self.current_item.setText(i, self.property_lineedits[i].text())

    def increment_last_alpha(self, s):
        last_alpha = s[-1]
        if last_alpha in string.ascii_uppercase:
            index = string.ascii_uppercase.index(last_alpha)
            if index + 1 < len(string.ascii_uppercase):
                if index == 7 or index == 14:
                    s = s[:-1] + string.ascii_uppercase[index + 2]
                else:
                    s = s[:-1] + string.ascii_uppercase[index + 1]
            else:
                self.show_alert('Z 이상의 첨자 품번을 생성할 수 없습니다.')
        # This part of the code checks if the last character of a string is a digit
        elif last_alpha.isdigit():
        # If it is a digit, then 'A' is added to the end of the string
            s += 'A'     
                                
        return s
    
    def show_alert(self, message):
        alert = QMessageBox()
        alert.setWindowTitle("Alert")
        alert.setText(message)
        alert.exec()

    def find_nodes_by_name(self, parent_item, name):
        result = []

        if parent_item is None:
            parent_item = self.qtree.invisibleRootItem()

        for i in range(parent_item.childCount()):
            child_item = parent_item.child(i)
            if child_item.text(0) == name:
                result.append(child_item)
            result.extend(self.find_nodes_by_name(child_item, name))

        return result
        
    def change_node_name(self, old_name, new_name):
        # Find all items by old_name
        items = self.find_nodes_by_name(None, old_name)

        for item in items:
            # Update the item's name
            item.setText(0, new_name)

            # Change the background color to red
            self.qtree.set_item_background_color(item, QtGui.QColor("red"))

            # Change parent node names and background color
            self.old_names.append(old_name)
            self.updated_names.append(new_name)
            self.change_node_and_ancestors(item.parent())
                    
    def change_node_and_ancestors(self, item): 
        if (item is None) or (item.text(0) in self.updated_names):
            return

        # Change the node name
        old_name = item.text(0)
        if ('620203-' not in old_name) and ('620205-' not in old_name):
            new_name = self.increment_last_alpha(old_name)
            self.old_names.append(old_name)
            self.updated_names.append(new_name)
        else:
            new_name = old_name

        # Update the item's name
        item.setText(0, new_name)
    
        # Change the background color to red
        self.qtree.set_item_background_color(item, QtGui.QColor("red"))

        # Move to the next parent
        self.change_node_and_ancestors(item.parent())  
 

    # def update_tree(self):
    #     self.qtree.clear()
    #     self.clickTreeBtn()

    # def update_table(self):
    #     self.table.clear()
    #     self.initTableWidget(2)  

    def on_change_name_button_clicked(self):
        try:
            old_name = self.old_name_edit.text()
            new_name = self.new_name_edit.text()
            self.change_node_name(old_name, new_name)
        except Exception as e:
            self.show_alert(str(e))
    

    def clickOpenBtn(self):
        file_path, ext = QFileDialog.getOpenFileName(self, '파일 열기', os.getcwd(), 'excel file (*.xls *.xlsx)')
        if file_path:
            self.df_list = self.loadData(file_path)
            self.clickTreeBtn()
            self.initTableWidget(id=2)
            
                       
    def on_search_button_clicked(self):
        text = self.line_edit.text()
        self.qtree.selectionModel().clear()
        self.qtree.do_search(text)
  
    def clickTreeBtn(self):
        TreeWidget = self.qtree
                
        font = QFont()
        font.setPointSize(12)  
        TreeWidget.setFont(font) 
        
        
        TreeWidget.setColumnCount(5)
        TreeWidget.setHeaderLabels(["ITM", "PREFIX", "ITM_DESC", "QTY", "UOM", "SRC", "PROC", "THREAD"] )
        TreeWidget.setAlternatingRowColors(True)
        TreeWidget.header().setSectionResizeMode(QHeaderView.Interactive)
        
        self.create_tree(self.df_list[2], TreeWidget)

        TreeWidget.expandAll()

        for i in range(5):
            TreeWidget.resizeColumnToContents(i)
        

               
    def create_tree(self, df, tree_widget):
        parent_dict = {}

        for index, row in df.iterrows():
            itm = row['ITM']
            parent_itm = row['PARENT']

            itm_data = [itm, row['PREFIX'], row['ITM_DESC'], str(row['QTY']),
                        row['UOM'], row['SRC'], row['PROC'], row['THREAD']]

            if parent_itm not in parent_dict:
                parent_dict[parent_itm] = QTreeWidgetItem(tree_widget, [parent_itm])

            itm_item = QTreeWidgetItem(parent_dict[parent_itm], itm_data)
            parent_dict[itm] = itm_item  
            
    def loadData(self, file_name):
        df_list = []        
        with pd.ExcelFile(file_name) as wb:            
            for i, sn in enumerate(wb.sheet_names):              
                try:
                    df = pd.read_excel(wb, sheet_name=sn)
                    if sn == "ECO_BOM": #ECO_BOM시트만 OLD BOM정리
                        df= df.iloc[3:,:11]
                        df.replace('\n','', regex=True, inplace=True)
                        df= df.rename(columns = df.iloc[0])
                        df= df.drop(df.index[0])
                        df.dropna(axis=0, how='all', inplace=True)
                        df= df
                        df= df.reset_index(drop=True)
                        
                except Exception as e:
                    print('File read error:', e)
                else:
                    df = df.fillna(0)
                    df.name = sn
                    df_list.append(df)
        return df_list

    def initTableWidget(self, id):
        # 테이블 위젯 값 쓰기
        self.table.clear()
        
        
        font = QFont()
        font.setPointSize(10)
        self.table.setFont(font) 
        
        
        # select dataframe
        df = self.df_list[id]; 

        # Add empty rows where necessary
        # df = self.insert_empty_rows(df)

        self.table.verticalHeader().setVisible(False)

        # table write        
        col = len(df.keys())
        self.table.setColumnCount(col)
        self.table.setHorizontalHeaderLabels(df.keys())
    
        row = len(df.index)
        self.table.setRowCount(row)
        self.writeTableWidget(df, row, col)





    def writeTableWidget(self, df, row, col):
        tree_row_height = 11
        TopRow = []

        table_row = 0

        for i in range(row): 
            item_row = df.iloc[i]

            if item_row['PARENT'] not in df['ITM'].values and item_row['PARENT'] not in TopRow: 
                TopRow.append(item_row['PARENT'])
                
                self.table.insertRow(table_row)
                self.table.setRowHeight(table_row, tree_row_height)

                table_row += 1

            for c in range(col):
                item = QTableWidgetItem(str(df.iloc[i][c]))
                self.table.setItem(table_row, c, item)
                self.table.setRowHeight(table_row, tree_row_height)

            table_row += 1
            
        self.table.resizeColumnsToContents()
    

 
    def save_dataframe_to_excel(self, df):

        options = QFileDialog.Options()
        options |= QFileDialog.DontUseNativeDialog
        file_path, _ = QFileDialog.getSaveFileName(self, "Save Excel File", "",
                                                   "Excel Files (*.xlsx);;All Files (*)", options=options)
        if file_path:

            # V5
            df = df[~df['ITM'].isin(self.deleted_items)]

            # Make the lists the same length as the dataframe
            old_names = list(OrderedDict.fromkeys(self.old_names))
            old_names.extend([None]*(len(df)-len(old_names)))
            new_names = list(OrderedDict.fromkeys(self.updated_names))
            new_names.extend([None]*(len(df)-len(new_names)))

            # Add Old Items and New Items Column
            df.insert(10, 'Old Items', old_names)
            df.insert(11, 'New Items', new_names)
            
            df.to_excel(file_path, index=False)
            
            # Load workbook
            wb = openpyxl.load_workbook(file_path)

            # Select sheet
            ws = wb.active

            # Define cell color
            fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type = "solid")


            for column in ('K', 'L'):  # Columns 'K' and 'L'
                for cell in ws[column]:
                    cell.fill = fill


            # Adjust width of all columns to fit the font size
            for column_cells in ws.columns:
                length = max(len(str(cell.value)) for cell in column_cells)
                ws.column_dimensions[column_cells[0].column_letter].width = length + 2

            # Save the modified Excel file
            wb.save(file_path)


if __name__ == '__main__':
    app = QApplication(sys.argv)
    w = myWindow()
    
    screen = app.primaryScreen()
    screen_geometry = screen.geometry()
    screen_width = screen_geometry.width()
    screen_height = screen_geometry.height()
    margin = 30 
    w.setGeometry(margin, margin, screen_width - 2 * margin, screen_height - 2 * margin)
    
    w.show()
    sys.exit(app.exec())   
